"""Excel Parser Executor - Parse Excel (.xlsx) files into structured row data."""

import os
import time
from datetime import date, datetime
from typing import Any, List

import structlog

from app.orchestrator.node_executors.base import ExecutionContext, NodeExecutionData

logger = structlog.get_logger()

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelParserExecutor:
    """
    Parse Excel (.xlsx) files with intelligent defaults.

    Features:
    - Read specific sheets or default to active sheet
    - Header row detection
    - Automatic type preservation (int, float, datetime, str)
    - Skip rows and row limits
    - Returns all sheet names for discoverability
    """

    MAX_ROWS = 100_000
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
    ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

    async def execute(
        self, data: NodeExecutionData, context: ExecutionContext
    ) -> dict[str, Any]:
        """Parse an Excel file and return structured row data."""
        if not HAS_OPENPYXL:
            return self._error_result(
                "openpyxl is required to parse Excel files. "
                "Install it with: pip install openpyxl"
            )

        start_time = time.monotonic()

        excel_input = data.inputs.get("excel_input")
        sheet_name = data.inputs.get("sheet_name")
        has_header = data.inputs.get("has_header", True)
        skip_rows = data.inputs.get("skip_rows", 0)
        max_rows = min(data.inputs.get("max_rows", self.MAX_ROWS), self.MAX_ROWS)

        # Validate input
        if not excel_input or not isinstance(excel_input, str):
            return self._error_result(
                "excel_input is required and must be a file path string"
            )

        if not os.path.isfile(excel_input):
            return self._error_result(f"File not found: {excel_input}")

        _, ext = os.path.splitext(excel_input)
        if ext.lower() not in self.ALLOWED_EXTENSIONS:
            return self._error_result(
                f"Unsupported file extension: {ext}. "
                f"Supported: {', '.join(sorted(self.ALLOWED_EXTENSIONS))}"
            )

        file_size = os.path.getsize(excel_input)
        if file_size > self.MAX_FILE_SIZE:
            return self._error_result(
                f"File too large: {file_size} bytes "
                f"(max {self.MAX_FILE_SIZE // (1024 * 1024)} MB)"
            )

        if file_size == 0:
            return self._error_result("File is empty")

        logger.info(
            "excel_parser_start",
            file_path=excel_input,
            file_size=file_size,
            sheet_name=sheet_name,
            has_header=has_header,
            skip_rows=skip_rows,
            max_rows=max_rows,
            node_id=data.node_id,
            workflow_id=context.workflow_id,
        )

        try:
            wb = openpyxl.load_workbook(
                excel_input, read_only=True, data_only=True
            )
        except Exception as e:
            return self._error_result(f"Failed to open Excel file: {e}")

        try:
            all_sheet_names: List[str] = wb.sheetnames

            if not all_sheet_names:
                return self._error_result("Workbook contains no sheets")

            if sheet_name:
                if sheet_name not in all_sheet_names:
                    return self._error_result(
                        f"Sheet '{sheet_name}' not found. "
                        f"Available sheets: {', '.join(all_sheet_names)}"
                    )
                ws = wb[sheet_name]
            else:
                ws = wb.active
                if ws is None:
                    ws = wb[all_sheet_names[0]]

            rows: List[dict[str, Any]] = []
            columns: List[str] = []
            row_iter = ws.iter_rows()

            skipped = 0
            while skipped < skip_rows:
                try:
                    next(row_iter)
                    skipped += 1
                except StopIteration:
                    break

            if has_header:
                try:
                    header_row = next(row_iter)
                    columns = self._extract_headers(header_row)
                except StopIteration:
                    return {
                        "rows": [],
                        "row_count": 0,
                        "columns": [],
                        "sheet_names": all_sheet_names,
                        "success": True,
                        "error": None,
                    }

            for row in row_iter:
                if len(rows) >= max_rows:
                    break

                cell_values = [self._convert_cell(cell.value) for cell in row]

                if all(v is None for v in cell_values):
                    continue

                if not columns:
                    columns = [
                        f"column_{i + 1}" for i in range(len(cell_values))
                    ]

                row_dict: dict[str, Any] = {}
                for idx, col_name in enumerate(columns):
                    if idx < len(cell_values):
                        row_dict[col_name] = cell_values[idx]
                    else:
                        row_dict[col_name] = None

                rows.append(row_dict)

            elapsed_ms = self._elapsed_ms(start_time)

            logger.info(
                "excel_parser_complete",
                row_count=len(rows),
                column_count=len(columns),
                sheet_used=ws.title,
                elapsed_ms=elapsed_ms,
                node_id=data.node_id,
            )

            return {
                "rows": rows,
                "row_count": len(rows),
                "columns": columns,
                "sheet_names": all_sheet_names,
                "success": True,
                "error": None,
            }

        except Exception as e:
            logger.error(
                "excel_parser_error",
                error=str(e),
                error_type=type(e).__name__,
                node_id=data.node_id,
            )
            return self._error_result(f"Error parsing Excel file: {e}")
        finally:
            wb.close()

    def _extract_headers(self, header_row) -> List[str]:
        """Extract column names from a header row, deduplicating empties."""
        headers: List[str] = []
        seen: set[str] = set()

        for idx, cell in enumerate(header_row):
            raw = cell.value
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                name = f"column_{idx + 1}"
            else:
                name = str(raw).strip()

            original = name
            suffix = 2
            while name in seen:
                name = f"{original}_{suffix}"
                suffix += 1

            seen.add(name)
            headers.append(name)

        return headers

    def _convert_cell(self, value: Any) -> Any:
        """Convert a cell value to a JSON-serializable Python type."""
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, str):
            return value
        return str(value)

    def _elapsed_ms(self, start_time: float) -> float:
        return round((time.monotonic() - start_time) * 1000, 2)

    def _error_result(self, error_message: str) -> dict[str, Any]:
        return {
            "rows": [],
            "row_count": 0,
            "columns": [],
            "sheet_names": [],
            "success": False,
            "error": error_message,
        }
